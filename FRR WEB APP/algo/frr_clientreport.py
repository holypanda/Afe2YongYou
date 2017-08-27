from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import string



COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',  # 0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',  # 5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
    '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
    '00993300', '00993366', '00333399', '00333333', 'System Foreground', 'System Background'  # 60-64
)

yellow = PatternFill(bgColor=COLOR_INDEX[5], start_color=COLOR_INDEX[5], end_color=COLOR_INDEX[5], fill_type="solid")


def update_excel(input_filename, output_filename):




    def delete_row(target):
        column = list(string.ascii_uppercase)
        for j in column:
            ws["%s%s" % (j, str(target))] = None

    # UPDATE HEADING
    def update_heading():
        ws["L1"] = "Floating"
        ws["M1"] = "Total Equity"
        ws["N1"] = "Initial Margin"
        ws["O1"] = "Main Margin"
        ws["P1"] = "Excess/(Decifit)"
        ws["Q1"] = "Margin Call"
        ws["R1"] = "Ratio"

        ws["S1"] = "Margin call"
        ws["S1"].fill = yellow
        ws["T1"] = "Excess"
        ws["T1"].fill = yellow
        ws["U1"] = "Excess/(Margin Call)"
        ws["U1"].fill = yellow
        ws["V1"] = "Initial Margin (HKD)"
        ws["V1"].fill = yellow
        ws["W1"] = "Working"
        ws["X1"] = "FRR Ranking liab"
        ws["X1"].fill = yellow
        ws["Y1"] = "Working"
        ws["Z1"] = "Initial Margin"
        ws["Z1"].fill = yellow

        ws["AA1"] = "USD"
        ws["AA2"] = "JPY"
        ws["AA3"] = "EUR"
        ws["AA4"] = "GBP"
        ws["AA5"] = "AUD"
        ws["AA6"] = "SGD"
        ws["AA7"] = "MYR"
        ws["AA8"] = "CNY"

    def load_exchange_rate(file="exchange_rate.txt"):
        exchange_rate_list = []
        with open(file, "r") as file:
            for line in file:
                rate = line.split(",")[1].strip()
                exchange_rate_list.append(rate)
        USD = float(exchange_rate_list[0])
        JPY = float(exchange_rate_list[1])
        EUR = float(exchange_rate_list[2])
        GBP = float(exchange_rate_list[3])
        AUD = float(exchange_rate_list[4])
        SGD = float(exchange_rate_list[5])
        MYR = float(exchange_rate_list[6])
        CNY = float(exchange_rate_list[7])
        HKD = float(exchange_rate_list[8])
        return USD, JPY, EUR, GBP, AUD, SGD, MYR, CNY, HKD

    # CHANGE EXCHANGE RATE MONTHLY
    def set_interest_rate():
        USD, JPY, EUR, GBP, AUD, SGD, MYR, CNY, HKDd = load_exchange_rate()
        ws["AB1"] = USD  # USD
        ws["AB2"] = JPY  # JPY
        ws["AB3"] = EUR  # EUR
        ws["AB4"] = GBP  # GBP
        ws["AB5"] = AUD  # AUD
        ws["AB6"] = SGD  # SGD
        ws["AB7"] = MYR  # MYR
        ws["AB8"] = CNY  # CNY
        # CHANGE EXCHANGE RATE MONTHLY



        # USD_rate = 7.809  # USD
        # JPY_rate = 0.0706  # JPY
        # EUR_rate = 9.1615  # EUR
        # GBP_rate = 10.2305 # GBP
        # AUD_rate = 6.2261  # AUD
        # SGD_rate = 5.758  # SGD
        # MYR_rate = 1.8455  # MYR
        # CNY_rate = 1.1598  # CNY

        # UPDATE S-V 4 FORMULA

    target_row_list = []

    wb = load_workbook(input_filename)

    sheet_name = wb.sheetnames

    ws = wb[sheet_name[0]]  # sheet selector

    update_heading()

    set_interest_rate()

# write row
    currency = None
    for i in range(1,600):
        row = i
        check_heading = ws["A%s" % str(i)].value

        if check_heading == "Client type : Normal (CNY)":
            currency = "CNY"

        elif check_heading == "Client type : Normal (GBP)":
            currency = "GBP"

        elif check_heading == "Client type : Normal (EUR)":
            currency = "EUR"

        elif check_heading == "Client type : Normal (USD)":
            currency = "USD"

        elif check_heading == "Client type : Normal (MYR)":
            currency = "MYR"

        elif check_heading == "Client type : Normal (JPY)":
            currency = "JPY"

        elif check_heading == "Client type : Normal (SGD)":
            currency = "SGD"

        elif check_heading == "Client type : Normal (AUD)":
            currency = "AUD"

        elif check_heading == "Client type : Normal (HKD)":
            currency = "HKD"

# write row

        elif currency == "CNY":
            # print("CNY",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$8" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$8,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$8" % row

            # row += 1

        elif currency == "GBP":
            # print("GBP",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$4" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$4,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$4" % row

            # row += 1

        elif currency == "EUR":
            # print("EUR",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$3" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$3,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$3" % row

            # row += 1

        elif currency == "USD":
            # print("USD",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$1" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$1,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$1" % row
            # row += 1

        elif currency == "MYR":
            # print("MYR",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$7" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$7,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$7" % row

            # row += 1

        elif currency == "JPY":
            # print("JPY",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$2" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$2,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$2" % row

            # row += 1

        elif currency == "SGD":
            # print("SGD",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$6" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$6,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$6" % row

            # row += 1

        elif currency == "AUD":
            # print("AUD",row," ",i)

            ws["S%s" % row] = "=-Q%s*$AB$5" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s)*$AB$5,0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s*$AB$5" % row

            # row += 1

        elif currency == "HKD":
            # print("HKD",row," ",i)

            ws["S%s" % row] = "=-Q%s" % row
            ws["T%s" % row] = "=IF(M%s+O%s>0,(M%s+O%s),0)" % (row, row, row, row)
            ws["U%s" % row] = "=IF(S%s<0,S%s,T%s)" % (row, row, row)
            ws["V%s" % row] = "=-N%s" % row

            # row += 1

# deleting not necessary row
    for i in range(1,600):
        check_heading = ws["A%s" % str(i)].value
        if not str(check_heading).startswith("3") \
                and check_heading != "Unclear Amt." \
                and check_heading != "Client type : Normal (HKD)" \
                and check_heading != "Client type : Normal (AUD)" \
                and check_heading != "Client type : Normal (SGD)" \
                and check_heading != "Client type : Normal (JPY)" \
                and check_heading != "Client type : Normal (MYR)" \
                and check_heading != "Client type : Normal (USD)" \
                and check_heading != "Client type : Normal (EUR)" \
                and check_heading != "Client type : Normal (GBP)" \
                and check_heading != "Client type : Normal (CNY)" :

            target_row_list.append(i)
            # print(i)
            for target_row in target_row_list:
                delete_row(target_row)


    row = 4
    ws["W%s" % row] = "=SUMIF(A:A,A%s,U:U)" % row
    # ws["W4"].fill = yellow
    ws["X%s" % row] = "=IF(W%s<0,IF(A%s=A%s,0,W%s),0)" % (row,row,row+1,row)
    # ws["X4"].fill = yellow
    ws["Y%s" % row] = "=SUMIF(A:A,A%s,V:V)" % row
    # ws["Y4"].fill = yellow
    ws["Z%s" % row] = "=IF(A%s=A%s,0,Y%s)" % (row,row+1,row)
    # ws["Z4"].fill = yellow



    wb.save(output_filename)
    wb.close()

if __name__ == "__main__":
    # input_file = input("input: ")
    # output_file = input("output: ")
    input = "clientreport 0824.xlsx"
    output = "after_%s" % input

    update_excel(input, output)
