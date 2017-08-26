from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import string
import os
import time
from shutil import copyfile

COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',  # 0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',  # 5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
    '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
    '00993300', '00993366', '00333399', '00333333', 'System Foreground', 'System Background'  # 60-64
)
yellow = PatternFill(bgColor=COLOR_INDEX[5], start_color=COLOR_INDEX[5], end_color=COLOR_INDEX[5], fill_type="solid")


# 读取afe
def load_gl_from_afe(file, sheet=0):
    wb = load_workbook(file)
    sheet_name = wb.sheetnames

    ws = wb[sheet_name[sheet]]  # sheet selector
    row = ws.max_row

    return ws, row


# 加载afe gl form， return ws
def load_gl_from(file):
    wb = load_workbook(file)
    sheet_name = wb.sheetnames

    ws = wb[sheet_name[0]]  # sheet selector
    row = ws.max_row

    return ws, row


# 提取gl data
def fetch_data_gl(worksheet, row):
    columns = list(string.ascii_uppercase)
    data = []
    for column in columns:
        info = worksheet["%s%s" % (column, str(row))].value
        data.append(info)
    return data


# 处理 data retrun一个列表
def process_data(file):
    ws, row = load_gl_from_afe(file)
    all_gl_data = []
    process_gl_data = []
    for i in range(1, row):
        row_data = fetch_data_gl(ws, i)
        all_gl_data.append(row_data)
    # print(all_gl_data)
    print("market, currency, account_id, account_des, debit, credit, summary")
    for gl_data in all_gl_data:
        if gl_data[0] == "HKEX" \
                or gl_data[0] == "CBT" \
                or gl_data[0] == "NYM" \
                or gl_data[0] == "NYB" \
                or gl_data[0] == "CME" \
                or gl_data[0] == "LME" \
                or gl_data[0] == "CMX" \
                or gl_data[0] == "IPE" \
                or gl_data[0] == "BMD" \
                or gl_data[0] == "JCCH" \
                or gl_data[0] == "LIFFE" \
                or gl_data[0] == "SGX":
            market = gl_data[0]
            currency = gl_data[1]
            account_id = gl_data[2]
            account_descr = gl_data[3]
            debit = gl_data[4]
            credit = gl_data[5]
            summary = gl_data[6]
            # print(market, currency, account_id, account_descr, debit, credit, summary)
            process_gl_data.append([market, currency, account_id, account_descr, debit, credit, summary])
    for gl_data in process_gl_data:
        if gl_data[1] == None and gl_data[2] == None:
            index = process_gl_data.index(gl_data)
            del process_gl_data[index]

    print("\n")
    print("market, currency, account_id, account_des, debit, credit, summary")
    print(process_gl_data)
    return process_gl_data


# 开始一个excel，建立heading
def initialize_form():
    wb = workbook.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws[
        "A1"] = '''"null_$head,main_m_pk_glorgbook,main_m_pk_vouchertype,main_m_no,main_pk_prepared,main_m_prepareddate,m_explanation,m_accsubjcode,m_pk_currtype,m_debitamount,m_localdebitamount,m_price,m_debitquantity,m_creditquantity,m_creditamount,m_localcreditamount,m_fracdebitamount,m_fraccreditamount,m_checkno,m_checkdate,m_free1,m_free2,m_bankaccount,m_free3,m_checkstyle,m_excrate2,m_excrate1,ass_1,ass_2,ass_3,ass_4,ass_5,ass_6,ass_7,ass_8,ass_9"'''

    ws["B1"] = "主体账簿"
    ws["C1"] = "凭证类别"
    ws["D1"] = "凭证号"
    ws["E1"] = "制单人"
    ws["F1"] = "制单日期"
    ws["G1"] = "摘要"
    ws["H1"] = "科目编码"
    ws["I1"] = "币种"
    ws["J1"] = "原币借方金额"
    ws["K1"] = "本币借方金额"
    ws["L1"] = "单价"
    ws["M1"] = "借方数量"
    ws["N1"] = "贷方数量"
    ws["O1"] = "原币贷方金额"
    ws["P1"] = "本币贷方金额"
    ws["Q1"] = "辅币借方金额"
    ws["R1"] = "辅币贷方金额"
    ws["S1"] = "结算号"
    ws["T1"] = "结算日期"
    ws["U1"] = "核销号"
    ws["V1"] = "核销日期"
    ws["W1"] = "银行账户"
    ws["X1"] = "票据类型"
    ws["Y1"] = "结算方式"
    ws["Z1"] = "折本汇率"
    ws["AA1"] = "折辅汇率"
    ws["AB1"] = "辅助核算1"
    ws["AC1"] = "辅助核算2"
    ws["AD1"] = "辅助核算3"
    ws["AE1"] = "辅助核算4"
    ws["AF1"] = "辅助核算5"
    ws["AG1"] = "辅助核算6"
    ws["AH1"] = "辅助核算7"
    ws["AI1"] = "辅助核算8"
    ws["AJ1"] = "辅助核算9"
    return ws, wb


# 写入一行
def write_row(ws, row_id, market, currency, account_id, account_descr, debit, credit, summary, ID, date, rate):
    book = "406-0004"
    proof_type = "TV"
    proof_id = "%s" % ID
    creater = "余豐"
    create_date = "%s" % date
    exchange_rate = "%s" % rate
    debit_currency_in_HKD = debit * rate
    credit_currency_in_HKD = credit * rate
    row_position = row_id + 2
    ws["A%s" % row_position] = str(int(row_id))
    ws["B%s" % row_position] = str(book)
    ws["C%s" % row_position] = str(proof_type)
    ws["D%s" % row_position] = str(proof_id)
    ws["E%s" % row_position] = str(creater)
    ws["F%s" % row_position] = str(create_date)
    ws["G%s" % row_position] = str(summary)
    ws["H%s" % row_position] = str(account_id)
    ws["I%s" % row_position] = str(currency)
    ws["J%s" % row_position] = str(debit)
    ws["K%s" % row_position] = str(debit_currency_in_HKD)

    ws["O%s" % row_position] = str(credit)
    ws["P%s" % row_position] = str(credit_currency_in_HKD)

    ws["Z%s" % row_position] = str(exchange_rate)

    if summary == "COMM AS AT %s (%s)" % (date, market) \
            or summary == "EXCHANGE FEE AS AT %s (%s)" % (date, market) \
            or summary == "EXCHANGE FEE AS AT %s (HKEX)" % date \
            or summary == "LEVY AS AT %s (HKEX)" % date:
        if account_id == "602100901" \
                or account_id == "602109903" \
                or account_id == "642100602" \
                or account_id == "642100903" \
                or account_id == "642100601" \
                or account_id == "642100902" \
                or account_id.startswith("6"):
            ws["AB%s" % row_position] = "BUSINESS：FU"
            ws["AC%s" % row_position] = "部门档案：FU"


def write_ending_row(ws, row_id):
    row_position = row_id + 3
    ws[
        "A%s" % row_position] = '''"cashflow,m_flag,cashflowcurr,m_money,m_moneymain,m_moneyass,cashflowName,cashflowCode"'''
    ws["B%s" % row_position] = "方向"
    ws["C%s" % row_position] = "分析币种"
    ws["D%s" % row_position] = "原币"
    ws["E%s" % row_position] = "本币"
    ws["F%s" % row_position] = "辅币"
    ws["G%s" % row_position] = "现金流量名称"
    ws["H%s" % row_position] = "现金流量编码"


# 从afe转换到yonyou
def transform_data(process_gl_data, ID, date, selector):
    ws, wb = initialize_form()

    # 字典 链接两份表

    def load_exchange_rate(file="exchange_rate.txt"):
        exchange_rate_list = []
        with open(file, "r") as file:
            for line in file:
                rate = line.split(",")[1].strip()
                exchange_rate_list.append(rate)
        USD = float(exchange_rate_list[0])
        JPY = float(exchange_rate_list[1])
        EUR = float(exchange_rate_list[2])
        GBP = float(exchange_rate_list[3])
        AUD = float(exchange_rate_list[4])
        SGD = float(exchange_rate_list[5])
        MYR = float(exchange_rate_list[6])
        CNY = float(exchange_rate_list[7])
        HKD = float(exchange_rate_list[8])
        return USD, JPY, EUR, GBP, AUD, SGD, MYR, CNY, HKD

    USD, JPY, EUR, GBP, AUD, SGD, MYR, CNY, HKD = load_exchange_rate()

    rate_dict = {"USD": USD,
                 "JPY": JPY,
                 "EUR": EUR,
                 "GBP": GBP,
                 "AUD": AUD,
                 "SGD": SGD,
                 "MYR": MYR,
                 "CNY": CNY,
                 "HKD": HKD}
    currency_dict = {"USD": "美元",
                     "JPY": "日圓",
                     "EUR": "歐羅",
                     "GBP": "英鎊",
                     "AUD": "澳元",
                     "SGD": "新加坡元",
                     "MYR": "馬來西亞林吉特",
                     "CNY": "人民币",
                     "HKD": "港币"}
    account_id_dict = {
        # HKD HKEX
        "A/R-BROKER HKFE - (CLIENT A/C)": "102100401",
        "A/P-CUSTODIAN CLIENT (LOCAL)": "231100101",
        "BROKERAGE-LOCAL": "602100901",
        "CLEARING FEE": "642100903",
        "TRANSACTION LEVY": "642100602",
        "FLOATING A/P-CUSTODIAN CLIENT (LOCAL)": "231100102",
        "COMPANY TRADING AC": "122109901",
        # USD CBT
        "A/R BROKER (ADMIS) USD-CLIENT A/C": "10120060103",  # ADMIS
        "A/R BROKER (MAREX) USD-CLIENT A/C": "10120061003",  # MAREX
        "BROKERAGE (OVERSEAS) - USD": "602100901",
        "EXCHANGE FEE (OS) USD": "602109903",
        "COMMISSION EXPENSES - USD": "642100601",
        # USD NYM
        "AR-OVERSEAS BROKERS (MAREX) USD": "10120060103",  # ADMIS
        # USD LME
        "EXCHANGE FEE TO BROKER (OS) USD": "642100902",
        # USD COMEX
        "COMPANY TRADING AC - (OS) USD": "122109901",
        # JPY SGX
        "BROKERAGE (OVERSEAS) - JPY": "602100901",
        "COMMISSION EXPENSES - JPY": "642100601",
        # MYR BMD
        "A/R-BROKER (PHILLIP FUTURES) MYR": "10120061201",
        "BROKERAGE (OVERSEAS) MYR" : "602100901",
        "COMMISSION EXPENSES - MYR" : "642100601"}

    row_id = 0
    for gl_data in process_gl_data:
        if gl_data[1] == "%s" % selector:
            market = gl_data[0]
            currency = currency_dict["%s" % gl_data[1]]
            account_id = gl_data[2]
            account_descr = gl_data[3]
            debit = gl_data[4]
            credit = gl_data[5]
            summary = gl_data[6]

            yonyou_summary_dict = {
                # HKD HKEX
                "Broker Ctrl A/C for trading Fee": "EXCHANGE FEE AS AT %s (%s)" % (date, market),
                "Levy for C.A.": "LEVY AS AT %s (%s)" % (date, market),
                "Client Trading P/L for CA / Broker": "TRADING P/L AS AT %s (%s)" % (date, market),
                "Exchange Fee for C.A.": "EXCHANGE FEE AS AT %s (%s)" % (date, market),
                "Floating P/L difference(Client) for Broker Floating A/C": "FLOATING P/L AS AT %s (%s)" % (
                    date, market),
                "Floating P/L difference(Client) for Client Floating A/C": "FLOATING P/L AS AT %s (%s)" % (
                    date, market),
                "Commission": "COMM AS AT %s (%s)" % (date, market),
                "Exchange Fee": "EXCHANGE FEE AS AT %s (%s)" % (date, market),
                "Levy": "LEVY AS AT %s (%s)" % (date, market),
                "Broker Exchange Fee": "EXCHANGE FEE AS AT %s (%s)" % (date, market),
                "Broker Levy Fee": "LEVY AS AT %s (%s)" % (date, market),
                "Trading P/L (Client)": "TRADING P/L AS AT %s (%s)" % (date, market),
                # USD CBT


                "Broker Commission": "COMM AS AT %s (%s)" % (date, market),
                # USD NYM
            }
            yonyou_summary = yonyou_summary_dict["%s" % summary]

            rate = rate_dict["%s" % gl_data[1]]

            # FLOATING 要用client floating的account
            if summary == "Floating P/L difference(Client) for Client Floating A/C" and account_descr == "A/P-CUSTODIAN CLIENT (LOCAL)":
                account_descr = "FLOATING A/P-CUSTODIAN CLIENT (LOCAL)"

            # LME, CMX, IPE, LIFFE use MAREX
            if account_descr == "AR-OVERSEAS BROKERS (MAREX) USD" and market == "LME":
                account_descr = "A/R BROKER (MAREX) USD-CLIENT A/C"
            if account_descr == "AR-OVERSEAS BROKERS (MAREX) USD" and market == "CMX":
                account_descr = "A/R BROKER (MAREX) USD-CLIENT A/C"
            if account_descr == "AR-OVERSEAS BROKERS (MAREX) USD" and market == "IPE":
                account_descr = "A/R BROKER (MAREX) USD-CLIENT A/C"
            if account_descr == "AR-OVERSEAS BROKER (MAREX) GBP" and market == "LIFFE":
                account_descr = "A/R BROKER (MAREX) USD-CLIENT A/C"

            # SGX + = PHILLIP FUTURE
            if account_descr == "A/R BROKER (ADMIS) USD-CLIENT A/C" and market == "SGX" and currency == "人民币":
                print(currency)
                account_descr = "A/R-BROKER (PHILLIP FUTURES) MYR"
                print(account_id_dict[account_descr])

            # convert account id
            account_id = account_id_dict["%s" % account_descr]

            # 过滤条件 适用于HKD和CNY
            if market == "HKEX" and summary == "Broker Ctrl A/C for trading Fee":
                pass
            elif market == "HKEX" and summary == "Broker Exchange Fee":
                pass
            elif market == "HKEX" and summary == "Broker Levy Fee":
                pass
            elif market == "HKEX" and summary == "Broker Ctrl A/C for trading Fee":
                pass
            elif market == "HKEX" and summary == "Broker Exchange Fee":
                pass
            elif market == "HKEX" and summary == "Broker Levy Fee":
                pass
            else:
                write_row(ws, row_id, market, currency, account_id, account_descr, debit, credit, yonyou_summary, ID,
                          date, rate)
                row_id += 1

    write_ending_row(ws, row_id)
    current_path = os.getcwd()
    wb.save("%s/yongyou/yongyou_gl_%s.xlsx" % (current_path, selector))


def convert_data_to_xls_format(currency_list):
    os.system("start ../formatConvert/convertor.xlsm")
    input("Enter any key after VBA scipte to move file to desire directory.")
    for c in currency_list:
        copyfile("../formatConvert/new gl_%s.xls" % c, "../Converted/converted gl_%s.xls" % c)
        copyfile("../formatConvert/new gl_%s.xls" % c, "C:/Users/TEMP/Desktop/Documents/converted gl_%s.xls" % c)
    for c in currency_list:
        os.remove("../formatConvert/new gl_%s.xls" % c)



if __name__ == "__main__":

    file = "gl 0824.xlsx"
    currency_list = ["CNY", "HKD", "USD", "JPY", "MYR"]
    # currency_list = ["CNY", "JPY", "MYR"]
    starting_ID = 211
    date = "2017-08-24"

    process_gl_data = process_data(file)
    for c in currency_list:
        transform_data(process_gl_data, ID=str(starting_ID), date=date, selector=c)
        starting_ID += 1

    convert_data_to_xls_format(currency_list)

    print('Finish!')





    # file = "gl 0823.xlsx"
    # process_gl_data = process_data(file)
    # currency = "CNY"
    # starting_ID = 168
    # date = "2017-08-18"
    # transform_data(process_gl_data, ID=str(starting_ID), date=date, selector=currency)
    # print('Finish!')



    # # ---------------------------------------------------------------------------------------- # #


    #
    #
    # file = input('file is (*.xlsx):')
    # currency = input('Currency(CNY,MYR):').upper()
    # currency_list = currency.split(",")
    # starting_ID = int(input('starting ID is: (100)'))
    # date = input('date is: (2017-08-16)')
    #
    # process_gl_data = process_data(file)
    # for c in currency_list:
    #     transform_data(process_gl_data, ID=str(starting_ID), date=date, selector=c)
    #     starting_ID += 1
